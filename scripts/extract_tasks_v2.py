import openpyxl
import json

path = "/Users/liuluheng/Library/CloudStorage/OneDrive-个人/Workfiles/05.项目运营/2026年度工作计划表.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb.active

results = []

def scan_band(start_row, end_row):
    for month_col_start in range(1, 16, 4): # Cols 1, 5, 9, 13 (1-indexed)
        month_name = sheet.cell(row=start_row, column=month_col_start).value
        if not month_name: continue
        
        for r in range(start_row + 1, end_row + 1):
            date_val = sheet.cell(row=r, column=month_col_start).value
            project = sheet.cell(row=r, column=month_col_start+1).value
            course = sheet.cell(row=r, column=month_col_start+2).value
            staff = sheet.cell(row=r, column=month_col_start+3).value
            
            if staff and "刘璐桁" in str(staff):
                results.append({
                    "month_header": month_name,
                    "date_str": str(date_val).strip() if date_val else None,
                    "project": str(project).strip() if project else None,
                    "course": str(course).strip() if course else None,
                    "staff": str(staff).strip()
                })

scan_band(2, 11)   # 1-4月
scan_band(12, 25)  # 5-8月
scan_band(26, 40)  # 9-12月 (extended to 40 just in case)

print(json.dumps(results, ensure_ascii=False, indent=2))
