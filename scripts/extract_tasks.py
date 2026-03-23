import openpyxl
import json

path = "/Users/liuluheng/Library/CloudStorage/OneDrive-个人/Workfiles/05.项目运营/2026年度工作计划表.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb["工作表1"]

# The months are likely arranged in columns.
# Based on peek:
# 1月: Col 0-3
# 2月: Col 4-7
# 3月: Col 8-11
# 4月: Col 12-15
# and so on.

results = []

# Headers are in row 2 (index 1)
# Data starts from row 3 (index 2)

for month_idx in range(0, 48, 4): # Assume 12 months * 4 columns = 48
    month_name = sheet.cell(row=2, column=month_idx+1).value
    if not month_name:
        continue
    
    for r in range(3, sheet.max_row + 1):
        date_val = sheet.cell(row=r, column=month_idx+1).value
        project = sheet.cell(row=r, column=month_idx+2).value
        course = sheet.cell(row=r, column=month_idx+3).value
        staff = sheet.cell(row=r, column=month_idx+4).value
        
        if staff and "刘璐桁" in str(staff):
            results.append({
                "month": month_name,
                "date_str": date_val,
                "project": project,
                "course": course,
                "staff": staff
            })

print(json.dumps(results, ensure_ascii=False, indent=2))
