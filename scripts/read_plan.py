import openpyxl
import json

try:
    path = "/Users/liuluheng/Library/CloudStorage/OneDrive-个人/Workfiles/05.项目运营/2026年度工作计划表.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    
    # Let's peek at the sheet names first
    print(f"Sheet names: {wb.sheetnames}")
    
    # We assume the first sheet or the active one
    sheet = wb.active
    results = []
    
    for row in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
        # Convert non-empty cells to string
        row_data = [str(c) if c is not None else "" for c in row]
        # Only add rows that have some content
        if any(row_data):
            results.append(" | ".join(row_data))
    
    print(json.dumps(results, ensure_ascii=False, indent=2))
except Exception as e:
    print(f"Error: {e}")
